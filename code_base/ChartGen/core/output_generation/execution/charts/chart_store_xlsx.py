"""
chart_store_xlsx.py
Download/upload round-trip for the Chart Store (workfile_config/chart_store.csv,
WorkfileState.chart_store_rows) -- the same simple full-replace pattern
stat_tags_xlsx.py uses for Stat Tags: download the current rows, edit in
Excel, upload to replace the whole list. No identity-merge logic -- a
Chart Store entry has no cached data of its own to preserve behind a
"deleted" flag, so a row absent from the uploaded file is simply gone.

Exports every CHART_STORE_FIELDNAMES column, including start_period/
end_period/metric_periods -- these aren't shown in the Charts sheet's own
Chart Store table (which shows only the human-relevant columns), but a
TimeSeries entry's cut can't be reproduced without them.
"""

from core.workfile.state.workfile_file import CHART_STORE_FIELDNAMES
from core.output_generation.execution.charts.chart_store import next_chart_store_id

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_chart_store_xlsx(chart_store_rows: list, output_path) -> str:
    """Write every Chart Store row (all CHART_STORE_FIELDNAMES columns) to a formatted .xlsx."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the Chart Store xlsx.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart Store"

    NAVY = "071A34"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    body_fill = PatternFill("solid", fgColor=LIGHT_GREY)
    body_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "chart_store_id": 12, "base_chart_name": 22, "cache_file": 16,
        "populations": 24, "start_period": 12, "end_period": 12,
        "metric_periods": 16, "width_emu": 12, "height_emu": 12,
        "tweaks": 30, "description": 40,
    }

    for col_idx, col_name in enumerate(CHART_STORE_FIELDNAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for excel_row, row in enumerate(chart_store_rows, start=2):
        for col_idx, col_name in enumerate(CHART_STORE_FIELDNAMES, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row.get(col_name, ""))
            cell.fill = body_fill
            cell.font = body_font
            cell.border = border
            cell.alignment = left_align if col_name in ("tweaks", "description") else centre_align
        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path


def read_chart_store_xlsx(path) -> list:
    """
    Read the Chart Store .xlsx (path or file-like buffer) and return a list
    of CHART_STORE_FIELDNAMES row dicts, one per non-empty data row -- the
    full replacement for WorkfileState.chart_store_rows (caller's
    responsibility; this function only parses).

    A row left with a blank "chart_store_id" is issued a fresh one here
    (via next_chart_store_id against the passed-in settings) rather than
    staying blank -- mirrors read_stat_tags_xlsx's own rule.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read the Chart Store xlsx.")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in excel_row):
            continue  # skip empty rows
        row = {}
        for col_idx, col_name in enumerate(CHART_STORE_FIELDNAMES):
            v = excel_row[col_idx] if col_idx < len(excel_row) else None
            row[col_name] = str(v).strip() if v is not None else ""
        rows.append(row)
    return rows


def assign_missing_chart_store_ids(rows: list, settings: dict) -> list:
    """
    Issue a fresh chart_store_id (next_chart_store_id) for any row read
    back with a blank "chart_store_id" column -- see read_chart_store_xlsx.
    Mutates settings' counter in place; caller marks the workfile dirty as
    usual.

    Passes every non-blank id already in `rows` to next_chart_store_id as
    existing_ids, so the counter resyncs against them before issuing a
    new one -- a row uploaded with its own id already filled in (the
    common case: re-uploading a previously-downloaded chart_store.xlsx)
    never advances the counter itself, so without this a fresh id issued
    afterwards could collide with one of those. The running set is
    updated as each new id is issued too, so two blank rows in the same
    upload can't collide with each other either.
    """
    existing_ids = {
        str(row.get("chart_store_id", "") or "").strip()
        for row in rows
        if str(row.get("chart_store_id", "") or "").strip()
    }
    for row in rows:
        if not str(row.get("chart_store_id", "") or "").strip():
            new_id = next_chart_store_id(settings, existing_ids)
            row["chart_store_id"] = new_id
            existing_ids.add(new_id)
    return rows
