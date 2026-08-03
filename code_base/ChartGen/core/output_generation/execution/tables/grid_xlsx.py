"""
grid_xlsx.py
Download/upload round-trip for a single Output Table's grid
(workfile_config/output_tables/{table_id}.csv, WorkfileState.output_tables)
-- the same full-replace pattern the Running Order and Stat Tags xlsx pairs
use (running_order/xlsx_writer.py+xlsx_reader.py; text/stat_tags_xlsx.py):
download the current grid, edit in Excel, upload replaces the whole grid.

Unlike those two, the sheet mirrors the grid's own spreadsheet shape
directly rather than a flat one-row-per-record table -- row 1 holds the
table_id (A1) and column widths; column A holds row heights; the rest is
content. Content cells get a dropdown of the workfile's current Stat Tag
ids ("[T3]" style), the same hidden-list-sheet pattern the Running Order's
period columns use (Architecture Decision 12) -- free text is still
accepted alongside the dropdown, matching that same precedent (no
showErrorMessage/errorStyle is set, so Excel doesn't reject a typed value
outside the list).
"""

from core.output_generation.execution.tables.grid_store import col_key

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_output_table_xlsx(grid_rows: list, stat_tag_rows: list, output_path) -> str:
    """
    Write a grid to a formatted .xlsx. stat_tag_rows (WorkfileState.
    text_stats_rows) supplies the dropdown options for content cells --
    every current tag, shown as its literal template text ("[T3]").
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the Output Table xlsx.")

    n_rows = max(0, len(grid_rows) - 1)
    n_cols = max(0, len(grid_rows[0]) - 1) if grid_rows else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grid"

    NAVY = "071A34"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    body_fill = PatternFill("solid", fgColor=LIGHT_GREY)
    body_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(n_rows + 1):
        for c in range(n_cols + 1):
            value = grid_rows[r].get(col_key(c), "") if r < len(grid_rows) else ""
            cell = ws.cell(row=r + 1, column=c + 1, value=value)
            cell.border = border
            cell.alignment = centre_align
            if r == 0 or c == 0:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = body_fill
                cell.font = body_font
            ws.column_dimensions[get_column_letter(c + 1)].width = 18
        ws.row_dimensions[r + 1].height = 20

    # --- Hidden list sheet: available Stat Tag ids for the content-cell dropdown ---
    tag_options = [f"[{row.get('tag', '')}]" for row in (stat_tag_rows or []) if row.get("tag", "")]
    if tag_options:
        tag_list_ws = wb.create_sheet("_stat_tag_list")
        tag_list_ws.sheet_state = "hidden"
        for opt_row, option in enumerate(tag_options, start=1):
            tag_list_ws.cell(row=opt_row, column=1, value=option)

        tag_dv = DataValidation(
            type="list",
            formula1=f"'_stat_tag_list'!$A$1:$A${len(tag_options)}",
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(tag_dv)
        # Content cells only -- rows 2..n_rows+1, cols 2..n_cols+1.
        for r in range(2, n_rows + 2):
            for c in range(2, n_cols + 2):
                tag_dv.add(ws.cell(row=r, column=c))

    ws.freeze_panes = "B2"

    wb.save(output_path)
    return output_path


def _cell_text(cell) -> str:
    """
    The cell's own displayed text, not its raw stored value. Excel's own
    behaviour: typing "5%" auto-applies a Percentage number format and
    stores the underlying value as the raw float 0.05 -- the "%" is
    purely display formatting, never part of the stored value itself.
    cell.value alone returns 0.05, and str(0.05) is exactly "0.05" --
    the literal bug this exists to close. Detected via cell.number_format
    containing "%"; the value is multiplied back up by 100 and the "%"
    appended, matching what Excel itself actually displays, not left as
    the underlying float. round() to a sane number of places first --
    floating point on an already-multiplied value can otherwise produce
    something like "5.000000000000001%".
    """
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (int, float)) and "%" in (cell.number_format or ""):
        pct = round(v * 100, 10)
        if pct == int(pct):
            pct = int(pct)
        return f"{pct}%"
    return str(v).strip()


def read_output_table_xlsx(path) -> list:
    """
    Read a grid .xlsx (path or file-like buffer) and return the full-
    replace grid -- a list of {"c0": ..., "c1": ..., ...} row dicts, one
    per used row, sized from the sheet's own used range (a user adding or
    removing rows/columns in Excel is honoured, not clamped back to the
    original dimensions). No validation of edited values here -- callers
    run grid_store.validate_grid() themselves, same as the on-screen
    Update button.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read the Output Table xlsx.")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Grid"] if "Grid" in wb.sheetnames else wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    rows = []
    for r in range(1, max_row + 1):
        row = {}
        for c in range(1, max_col + 1):
            row[col_key(c - 1)] = _cell_text(ws.cell(row=r, column=c))
        rows.append(row)
    return rows
