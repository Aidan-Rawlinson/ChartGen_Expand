"""
xlsx_writer.py
Writes Running Order rows to a formatted .xlsx file, with dropdown
validation on function, chart_type_ref, and enabled, and colour-coding by
row type.
"""

from core.output_generation.definition.running_order.schema import (
    COLUMNS, ALL_FUNCTIONS, SCOPE_VALUES, STRUCTURAL_FUNCTIONS,
)
from core.output_generation.definition.running_order.dialog_support import (
    get_valid_chart_refs_for_cache_file,
)

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_xlsx(rows: list[dict], output_path: str,
               manifest: dict = None, periods_by_cache_file: dict = None):
    """
    Write Running Order rows to a formatted .xlsx file with dropdown
    validation on function, chart_type_ref, enabled, and (for TimeSeries
    rows) start_period/end_period.

    periods_by_cache_file: {cache_file: [(period_id, period_label), ...]},
    in the shape's own trusted-chronological order — built by the caller
    (running_order_tab.py) via cache_reader.periods_for_cache_file, since
    it requires reading actual cached data, not just the manifest. Only
    cache files actually referenced by an insert_chart row need an entry.

    start_period/end_period are written and validated as
    "period_label(period_id)" rather than a bare id — the id alone tells
    the user nothing, and a bare label (e.g. "Jan 24") risks Excel silently
    reinterpreting the cell as a date. The parenthesised id also makes the
    value self-describing on read-back (xlsx_reader.py extracts it), so
    the cell keeps working even if periods are reordered between exports.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the Running Order xlsx.")
    periods_by_cache_file = periods_by_cache_file or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Running Order"

    # --- Hidden list sheet for start_period/end_period dropdowns ---
    # Excel's inline list formula1 (used for function/chart_type_ref/etc.)
    # is capped at 255 characters — fine for a handful of options, not for
    # a chart's full period history. Each distinct cache_file's period
    # options get their own column here (consecutive — column 1 for the
    # first cache_file encountered, column 2 for the next, and so on);
    # start_period and end_period both validate against the same column
    # for a given cache_file, since they share one option list.
    period_list_ws = wb.create_sheet("_period_lists")
    period_list_ws.sheet_state = "hidden"

    period_dv_by_cache_file = {}
    list_col_idx = 1
    for cache_file, period_pairs in periods_by_cache_file.items():
        if not period_pairs:
            continue
        options = [f"{label}({pid})" for pid, label in period_pairs]
        col_letter = get_column_letter(list_col_idx)
        for opt_row, option in enumerate(options, start=1):
            period_list_ws.cell(row=opt_row, column=list_col_idx, value=option)
        list_col_idx += 1

        period_dv = DataValidation(
            type="list",
            formula1=f"'_period_lists'!${col_letter}$1:${col_letter}${len(options)}",
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(period_dv)
        period_dv_by_cache_file[cache_file] = period_dv

    # --- Styles ---
    NAVY = "071A34"
    CRIMSON = "C12958"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"
    DISABLED_GREY = "AAAAAA"
    CHART_GREEN = "E8F5E9"
    PICTURE_TEAL = "E0F7FA"
    EXCEL_PURPLE = "F3E5F5"
    BATCH_ORANGE = "FFF3E0"
    STRUCTURAL_BLUE = "E3F2FD"
    POPULATIONS_AMBER = "FFF8E1"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    structural_fill = PatternFill("solid", fgColor=STRUCTURAL_BLUE)
    chart_fill = PatternFill("solid", fgColor=CHART_GREEN)
    disabled_font = Font(color=DISABLED_GREY, size=10)
    normal_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Column widths ---
    col_widths = {
        "row_id":        6,
        "enabled":       8,
        "scope":         13,
        "function":      22,
        "slide_index":   11,
        "chart_type_ref":22,
        "cache_file":    30,
        "populations":   30,
        "start_period":  16,
        "end_period":    16,
        "metric_periods":40,
        "image_path":    36,
        "excel_path":    36,
        "export_range":  18,
        "driver_range":  18,
        "left_emu":      12,
        "top_emu":       12,
        "width_emu":     12,
        "height_emu":    12,
        "notes":         40,
    }

    # --- Header row ---
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # --- Data validation: function dropdown (applies to entire function column) ---
    func_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ALL_FUNCTIONS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(func_dv)

    enabled_dv = DataValidation(
        type="list",
        formula1='"1,0"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(enabled_dv)

    scope_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(SCOPE_VALUES) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(scope_dv)

    # --- Data rows ---
    for data_row_idx, row in enumerate(rows, start=2):
        excel_row = data_row_idx
        is_enabled = str(row.get("enabled", "1")) == "1"
        func  = row.get("function", "")
        scope = str(row.get("scope", "normal")).strip()
        is_structural = func in STRUCTURAL_FUNCTIONS

        row_cache_file = str(row.get("cache_file") or "").strip()
        if row_cache_file.lower() == "none":
            row_cache_file = ""
        row_periods = periods_by_cache_file.get(row_cache_file, []) if func == "insert_chart" else []
        # Canonical storage is period_id alone; the cell shows
        # "period_label(period_id)" so the user sees a real label instead
        # of an id, and so the id survives a round-trip unambiguously.
        # A stored id no longer among the current periods (stale/legacy)
        # is shown as-is rather than dropped silently.
        id_to_display = {pid: f"{label}({pid})" for pid, label in row_periods}

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if value == "" or value is None:
                value = ""
            if col_name in ("start_period", "end_period") and value:
                value = id_to_display.get(str(value).strip(), value)
            elif col_name == "metric_periods" and value:
                # metric_periods can hold several ids in one cell ('^'-joined).
                # The per-row dropdown (below) still offers a single period
                # at a time — Excel's list validation has no multi-select —
                # but a cell already holding more than one (Charts sheet, or
                # typed by hand) keeps working; each token gets the same
                # "label(id)" treatment for readability and round-trip safety.
                tokens = str(value).split("^")
                value = "^".join(id_to_display.get(t.strip(), t.strip()) for t in tokens if t.strip())
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.font = disabled_font if not is_enabled else normal_font

            if scope in ("batch_open", "batch_close"):
                cell.fill = PatternFill("solid", fgColor=BATCH_ORANGE)
            elif is_structural and func == "set_default_populations":
                cell.fill = PatternFill("solid", fgColor=POPULATIONS_AMBER)
            elif is_structural:
                cell.fill = structural_fill
            elif func == "insert_chart":
                cell.fill = chart_fill
            elif func == "insert_picture":
                cell.fill = PatternFill("solid", fgColor=PICTURE_TEAL)
            elif func in ("insert_from_excel",):
                cell.fill = PatternFill("solid", fgColor=EXCEL_PURPLE)
            else:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

            # Alignment
            if col_name in ("row_id", "enabled", "slide_index",
                            "left_emu", "top_emu", "width_emu", "height_emu"):
                cell.alignment = centre_align
            else:
                cell.alignment = left_align

        # Apply function dropdown
        func_col = COLUMNS.index("function") + 1
        func_dv.add(ws.cell(row=excel_row, column=func_col))

        # Apply enabled dropdown
        enabled_col = COLUMNS.index("enabled") + 1
        enabled_dv.add(ws.cell(row=excel_row, column=enabled_col))

        # Apply scope dropdown
        scope_col = COLUMNS.index("scope") + 1
        scope_dv.add(ws.cell(row=excel_row, column=scope_col))

        # Per-row chart_type_ref dropdown — constrained to the valid chart
        # refs for the row's data shape via the shared rule in dialog_support
        # (which itself falls back to all refs if the shape is unknown).
        if func == "insert_chart":
            cache_file = str(row.get("cache_file") or "").strip()
            if cache_file.lower() == "none":
                cache_file = ""
            converts_to_metrics = bool(str(row.get("metric_periods") or "").strip())
            chart_refs = get_valid_chart_refs_for_cache_file(cache_file, manifest or {}, converts_to_metrics)

            if chart_refs:
                ref_formula = '"' + ",".join(chart_refs) + '"'
                chart_dv = DataValidation(
                    type="list",
                    formula1=ref_formula,
                    allow_blank=True,
                    showDropDown=False,
                )
                ws.add_data_validation(chart_dv)
                ctr_col = COLUMNS.index("chart_type_ref") + 1
                chart_dv.add(ws.cell(row=excel_row, column=ctr_col))

            # Per-row start_period/end_period/metric_periods dropdown — only
            # for a TimeSeries cache_file with a known period list. All
            # three columns validate against the same hidden-sheet range
            # for this cache_file (built once, above), not a fresh list per
            # row. The dropdown itself always offers one period at a time —
            # metric_periods is the one column where more than one may be
            # wanted, and Excel's list validation is single-value only. A
            # cell already holding a '^'-delimited value (from the Charts
            # sheet's multi-select, or typed by hand) isn't blocked by the
            # dropdown being there; the dropdown just makes adding/replacing
            # a single value easy without needing to know a period_id.
            period_dv = period_dv_by_cache_file.get(row_cache_file)
            if period_dv is not None:
                for period_col_name in ("start_period", "end_period", "metric_periods"):
                    period_col = COLUMNS.index(period_col_name) + 1
                    period_dv.add(ws.cell(row=excel_row, column=period_col))

        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path
