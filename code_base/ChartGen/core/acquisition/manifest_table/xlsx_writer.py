"""
xlsx_writer.py
Writes the manifest table to a formatted .xlsx for user editing — the same
download/edit/upload pattern as the Running Order xlsx.

Deliberately excluded columns: chart_ref (renumbered on import, so exporting
it would only invite conflicts) and deleted (deleted rows are simply not
exported; deleting a row in Excel is how a user deletes a chart).

hex_id is exported — it is the round-trip identity. Rows that come back
with a hex_id are existing rows; rows without one are new. Users add a
chart by adding a row with just a URL.
"""

from core.workfile.state.workfile_file import MANIFEST_FIELDNAMES

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Export column order — MANIFEST_FIELDNAMES minus chart_ref and deleted.
EXPORT_COLUMNS = [c for c in MANIFEST_FIELDNAMES if c not in ("chart_ref", "deleted")]

# Blank formatted rows appended below the data for user-entered charts.
BLANK_INPUT_ROWS = 300


def write_manifest_xlsx(manifest_rows: list[dict], output_path) -> str:
    """Write non-deleted manifest rows to a formatted .xlsx file or file-like buffer."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the manifest xlsx.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart URLs"

    # --- Styles (matching the Running Order xlsx palette) ---
    NAVY = "071A34"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"
    SYSTEM_GREY = "808080"
    URL_GREEN = "E8F5E9"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    system_fill = PatternFill("solid", fgColor=LIGHT_GREY)
    system_font = Font(color=SYSTEM_GREY, size=10)
    url_fill = PatternFill("solid", fgColor=URL_GREEN)
    normal_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "hex_id":          10,
        "url":             70,
        "chart_title":     40,
        "database":        10,
        "project_id":      11,
        "service_id":      11,
        "year":            10,
        "shape_type":      24,
        "source":          14,
        "added_at":        22,
        "data_updated_at": 22,
    }

    # --- Header row ---
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # --- Data rows: non-deleted only ---
    live_rows = [r for r in manifest_rows if str(r.get("deleted", "0")) != "1"]
    for excel_row, row in enumerate(live_rows, start=2):
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row.get(col_name, ""))
            cell.border = border
            if col_name == "url":
                cell.fill = url_fill
                cell.font = normal_font
                cell.alignment = left_align
            else:
                cell.fill = system_fill
                cell.font = system_font
                cell.alignment = centre_align if col_name in (
                    "hex_id", "database", "project_id", "service_id", "year",
                ) else left_align
        ws.row_dimensions[excel_row].height = 18

    # --- Blank formatted input rows for user-entered charts ---
    first_blank = len(live_rows) + 2
    for excel_row in range(first_blank, first_blank + BLANK_INPUT_ROWS):
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value="")
            cell.border = border
            if col_name == "url":
                cell.fill = url_fill
                cell.font = normal_font
                cell.alignment = left_align
            else:
                cell.fill = system_fill
                cell.font = system_font
                cell.alignment = left_align
        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path
