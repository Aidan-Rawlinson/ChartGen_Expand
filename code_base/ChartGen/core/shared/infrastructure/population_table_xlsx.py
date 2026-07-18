"""
population_table_xlsx.py
Excel export/import round-trip for population-level tables (nhs_organisations,
any submissions_{year}_{project_id} table, submissions_timeseries_{project_id},
and any future table) — the same download/edit/upload pattern as the
manifest table (core.acquisition.manifest_table) and the Running Order.

Generic across any population table: column list is read from the table's
own rows (spine first, via SPINE_COLUMN_ORDER, then any Name() columns in
their existing order) rather than a fixed schema, matching the "identical
headers, no single fixed schema" convention population tables already
follow (see Architecture, Section 5).

Identity is unit_id, not a system-generated key like the manifest's hex_id —
unit_id originates from the API (or, for a manually-added row, from the
user). Import semantics:
- Row present with a unit_id matching an existing row -> existing row;
  fields updated to the file's values.
- Row present with a unit_id not matching any existing row -> new row.
- Row present with a blank unit_id -> skipped; there's nothing to anchor
  a population-table row to without one (contrast the manifest table,
  where the URL itself identifies a new row).
- Known unit_id absent from the file -> row removed from the table. There
  is no soft-delete flag on population tables (unlike the manifest's
  deleted column) — removal is real. A soft_parents reference elsewhere
  pointing at a removed unit_id is left as-is; resolution already treats
  an unresolvable id as skipped (Functional Spec, Section 7.2), so this
  is tolerated rather than cleaned up.
"""

from core.shared.infrastructure.constants import SPINE_COLUMN_ORDER

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Blank formatted rows appended below the data for manually-added rows.
BLANK_INPUT_ROWS = 100


def _table_columns(rows: list) -> list:
    """Spine columns first, then any other (Name() peer-group) columns, in
    the order they appear on the row. Mirrors
    core.ui.common.formatting.population_table_columns — kept here too so
    this module has no dependency on the ui package (one-way dependency
    rule, Architecture Section 2)."""
    if not rows:
        return []
    present = list(rows[0].keys())
    spine = [c for c in SPINE_COLUMN_ORDER if c in present]
    rest = [c for c in present if c not in SPINE_COLUMN_ORDER]
    return spine + rest


def write_population_table_xlsx(table_name: str, rows: list[dict], output_path) -> str:
    """Write a population table to a formatted .xlsx file or file-like buffer."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the population table xlsx.")

    columns = _table_columns(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31] or "Population"

    NAVY = "071A34"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"
    EDIT_FILL_COLOUR = "FFF8E1"  # pale amber — everything here is user-editable

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    edit_fill = PatternFill("solid", fgColor=EDIT_FILL_COLOUR)
    normal_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "unit_id":      14,
        "unit_code":    16,
        "unit_name":    36,
        "soft_parents": 40,
    }

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for excel_row, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row.get(col_name, ""))
            cell.border = border
            cell.fill = edit_fill
            cell.font = normal_font
            cell.alignment = centre_align if col_name in ("unit_id", "unit_code") else left_align
        ws.row_dimensions[excel_row].height = 18

    first_blank = len(rows) + 2
    for excel_row in range(first_blank, first_blank + BLANK_INPUT_ROWS):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value="")
            cell.border = border
            cell.fill = edit_fill
            cell.font = normal_font
            cell.alignment = left_align
        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path


def read_population_table_xlsx(path) -> list[dict]:
    """
    Read an edited population-table .xlsx (path or file-like buffer) and
    return a list of dicts, one per non-empty data row, keyed by the
    header row's own column names.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read the population table xlsx.")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue  # skip empty rows
        record = {}
        for col_idx, col_name in enumerate(header):
            if not col_name:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            record[col_name] = "" if value is None else str(value).strip()
        rows.append(record)
    return rows


def apply_population_table_import(table_name: str, imported: list[dict], *, workfile_state) -> dict:
    """
    Apply read_population_table_xlsx output to workfile_state.tables[table_name]
    per the import semantics in the module docstring. Returns
    {"added": int, "updated": int, "deleted": int, "skipped_blank_id": int}.
    """
    existing_rows = workfile_state.tables.get(table_name, [])
    by_unit_id = {r["unit_id"]: r for r in existing_rows if r.get("unit_id")}
    columns = _table_columns(existing_rows) or (list(imported[0].keys()) if imported else [])

    seen_unit_ids = set()
    added = updated = skipped_blank_id = 0
    new_rows = []

    for imp in imported:
        unit_id = imp.get("unit_id", "").strip()
        if not unit_id:
            skipped_blank_id += 1
            continue

        existing = by_unit_id.get(unit_id)
        if existing is None:
            row = {col: imp.get(col, "") for col in columns}
            row["unit_id"] = unit_id
            new_rows.append(row)
            seen_unit_ids.add(unit_id)
            added += 1
            continue

        seen_unit_ids.add(unit_id)
        changed = False
        for col in columns:
            if col == "unit_id":
                continue
            new_value = imp.get(col, existing.get(col, ""))
            if new_value != existing.get(col, ""):
                existing[col] = new_value
                changed = True
        if changed:
            updated += 1

    kept_rows = [r for r in existing_rows if r.get("unit_id") in seen_unit_ids]
    deleted = len(existing_rows) - len(kept_rows)

    workfile_state.tables[table_name] = kept_rows + new_rows
    workfile_state.dirty = True

    return {"added": added, "updated": updated, "deleted": deleted,
             "skipped_blank_id": skipped_blank_id}
