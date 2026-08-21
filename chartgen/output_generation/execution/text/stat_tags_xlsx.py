"""
stat_tags_xlsx.py
Download/upload round-trip for the stat tags table (workfile_config/text_stats.csv,
WorkfileState.text_stats_rows) — the same simple full-replace pattern the
Running Order xlsx uses (running_order/xlsx_writer.py, xlsx_reader.py):
download the current rows, edit in Excel, upload to replace the whole list.
No identity-merge logic (contrast the manifest table's own hex_id-keyed
add/update/soft-delete round trip) — a stat tag has no cached data of its
own to preserve behind a "deleted" flag, so a row absent from the uploaded
file is simply gone, same as using the Text tab's own Delete button.

Exports every TEXT_STATS_FIELDNAMES column, including start_period/
end_period/metric_periods — these don't appear in the Text tab's own
on-screen table (which shows only the human-relevant columns), but a
TimeSeries tag's cut can't be reproduced without them, so the extract
would silently corrupt a TimeSeries tag if they were dropped.
"""

from chartgen.workfile.state.workfile_file import TEXT_STATS_FIELDNAMES
from chartgen.output_generation.execution.text.stat_tags import next_stat_tag

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_stat_tags_xlsx(stat_rows: list, output_path) -> str:
    """Write every stat tag row (all TEXT_STATS_FIELDNAMES columns) to a formatted .xlsx."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the stat tags xlsx.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stat Tags"

    NAVY = "071A34"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    body_fill = PatternFill("solid", fgColor=LIGHT_GREY)
    body_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "tag": 8, "hex_id": 10, "populations": 24,
        "start_period": 12, "end_period": 12, "metric_periods": 16,
        "reference_id": 12, "description": 40,
    }

    for col_idx, col_name in enumerate(TEXT_STATS_FIELDNAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for excel_row, row in enumerate(stat_rows, start=2):
        for col_idx, col_name in enumerate(TEXT_STATS_FIELDNAMES, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row.get(col_name, ""))
            cell.fill = body_fill
            cell.font = body_font
            cell.border = border
            cell.alignment = left_align if col_name == "description" else centre_align
        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path


def read_stat_tags_xlsx(path) -> list:
    """
    Read the stat tags .xlsx (path or file-like buffer) and return a list
    of TEXT_STATS_FIELDNAMES row dicts, one per non-empty data row — the
    full replacement for WorkfileState.text_stats_rows (caller's
    responsibility; this function only parses).

    A row left with a blank "tag" is issued a fresh one here (via
    next_stat_tag against the passed-in settings) rather than staying
    blank — a blank tag can never be matched to any [tag] in a template,
    so leaving it blank would just silently create a dead row. Any other
    blank column is left blank.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read the stat tags xlsx.")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in excel_row):
            continue  # skip empty rows
        row = {}
        for col_idx, col_name in enumerate(TEXT_STATS_FIELDNAMES):
            v = excel_row[col_idx] if col_idx < len(excel_row) else None
            row[col_name] = str(v).strip() if v is not None else ""
        rows.append(row)
    return rows


def assign_missing_tags(rows: list, settings: dict) -> list:
    """
    Issue a fresh tag (next_stat_tag) for any row read back with a blank
    "tag" column — see read_stat_tags_xlsx. Mutates settings' counter in
    place; caller marks the workfile dirty as usual.
    """
    for row in rows:
        if not str(row.get("tag", "") or "").strip():
            row["tag"] = next_stat_tag(settings)
    return rows
