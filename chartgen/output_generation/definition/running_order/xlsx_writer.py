"""
xlsx_writer.py
Writes Running Order rows to a formatted .xlsx file, with dropdown
validation on function, base_chart_name, and enabled, and colour-coding by
row type.
"""

from chartgen.output_generation.definition.running_order.schema import (
    COLUMNS, ALL_FUNCTIONS, SCOPE_VALUES, STRUCTURAL_FUNCTIONS,
)
from chartgen.output_generation.definition.running_order.dialog_support import (
    get_valid_chart_refs_for_cache_file,
)

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_xlsx(rows: list[dict], output_path: str,
               manifest: dict = None, periods_by_cache_file: dict = None,
               custom_chart_rows: list = None):
    """
    Write Running Order rows to a formatted .xlsx file with dropdown
    validation on function, base_chart_name, enabled, and (for TimeSeries
    rows) start_period/end_period.

    custom_chart_rows: this workfile's saved custom charts (WorkfileState.
    custom_chart_rows) — merged into the per-row base_chart_name dropdown
    for the row's own shape, the same as the Charts sheet and Running
    Order edit dialog. Omit or pass None/[] to offer built-in chart types
    only.

    periods_by_cache_file: {cache_file: [(period_id, period_label), ...]},
    in the shape's own trusted-chronological order — built by the caller
    (running_order_tab.py) via cache_reader.periods_for_cache_file, since
    it requires reading actual cached data, not just the manifest. Only
    cache files actually referenced by an insert_chart row need an entry.
    Used only to build the start_period/end_period/metric_periods dropdown
    option lists.

    start_period/end_period/metric_periods are written exactly as stored on
    the row. No derivation, no lookup, nothing recomputed here or on read.
    The numeric id is extracted only in cut_resolution.prepare_chart_cut.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the Running Order xlsx.")
    periods_by_cache_file = periods_by_cache_file or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Running Order"

    # --- Hidden list sheet for the three period dropdowns. Options are
    # "period_label(period_id)", matching the stored form exactly, so a
    # dropdown pick produces the same string the Charts sheet would store.
    #
    # A hidden sheet rather than an inline list because Excel caps inline
    # list validation at 255 characters, which a full period history
    # exceeds. One column per distinct cache_file, consecutive; all three
    # period columns validate against the same column. ---
    period_list_ws = wb.create_sheet("_period_lists")
    period_list_ws.sheet_state = "hidden"

    period_dv_by_cache_file = {}
    list_col_idx = 1
    for cache_file, period_pairs in periods_by_cache_file.items():
        if not period_pairs:
            continue
        options = [f"{label}({pid})" for pid, label in period_pairs]
        col_letter = get_column_letter(list_col_idx)
        for opt_row, option in enumerate(options, start=1):
            period_list_ws.cell(row=opt_row, column=list_col_idx, value=option)
        list_col_idx += 1

        period_dv = DataValidation(
            type="list",
            formula1=f"'_period_lists'!${col_letter}$1:${col_letter}${len(options)}",
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(period_dv)
        period_dv_by_cache_file[cache_file] = period_dv

    # --- Styles ---
    NAVY = "071A34"
    CRIMSON = "C12958"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"
    DISABLED_GREY = "AAAAAA"
    CHART_GREEN = "E8F5E9"
    PICTURE_TEAL = "E0F7FA"
    EXCEL_PURPLE = "F3E5F5"
    BATCH_ORANGE = "FFF3E0"
    STRUCTURAL_BLUE = "E3F2FD"
    POPULATIONS_AMBER = "FFF8E1"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    structural_fill = PatternFill("solid", fgColor=STRUCTURAL_BLUE)
    chart_fill = PatternFill("solid", fgColor=CHART_GREEN)
    disabled_font = Font(color=DISABLED_GREY, size=10)
    normal_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Column widths ---
    col_widths = {
        "row_id":        6,
        "enabled":       8,
        "scope":         13,
        "function":      22,
        "slide_index":   11,
        "base_chart_name":22,
        "cache_file":    30,
        "populations":   30,
        "start_period":  16,
        "end_period":    16,
        "metric_periods":40,
        "image_path":    36,
        "excel_path":    36,
        "export_range":  18,
        "driver_range":  18,
        "left_emu":      12,
        "top_emu":       12,
        "width_emu":     12,
        "height_emu":    12,
        "hyperlink_left":  14,
        "hyperlink_top":   14,
        "hyperlink_size":  14,
        "hyperlink_colour":14,
        "tweaks":        30,
        "notes":         40,
    }

    # --- Header row ---
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # --- Data validation: function dropdown (applies to entire function column) ---
    func_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ALL_FUNCTIONS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(func_dv)

    enabled_dv = DataValidation(
        type="list",
        formula1='"1,0"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(enabled_dv)

    scope_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(SCOPE_VALUES) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(scope_dv)

    # --- Data rows ---
    for data_row_idx, row in enumerate(rows, start=2):
        excel_row = data_row_idx
        is_enabled = str(row.get("enabled", "1")) == "1"
        func  = row.get("function", "")
        scope = str(row.get("scope", "normal")).strip()
        is_structural = func in STRUCTURAL_FUNCTIONS

        row_cache_file = str(row.get("cache_file") or "").strip()
        if row_cache_file.lower() == "none":
            row_cache_file = ""

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if value == "" or value is None:
                value = ""
            if col_name == "cache_file" and value:
                # The displayed value drops the ".json" suffix that is the
                # real cache dict key. xlsx_reader.py adds it back on read.
                # Row-level logic above still uses the untrimmed value; only
                # this cell's display
                # changes.
                stripped = str(value).strip()
                if stripped.lower().endswith(".json"):
                    stripped = stripped[:-5]
                value = stripped
            # start_period/end_period/metric_periods are written exactly
            # as stored -- see module docstring for why nothing is
            # derived or rewritten here.
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.font = disabled_font if not is_enabled else normal_font

            if scope in ("batch_open", "batch_close"):
                cell.fill = PatternFill("solid", fgColor=BATCH_ORANGE)
            elif is_structural and func == "set_default_populations":
                cell.fill = PatternFill("solid", fgColor=POPULATIONS_AMBER)
            elif is_structural:
                cell.fill = structural_fill
            elif func == "insert_chart":
                cell.fill = chart_fill
            elif func == "insert_picture":
                cell.fill = PatternFill("solid", fgColor=PICTURE_TEAL)
            elif func in ("insert_from_excel",):
                cell.fill = PatternFill("solid", fgColor=EXCEL_PURPLE)
            else:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

            # Alignment
            if col_name in ("row_id", "enabled", "slide_index",
                            "left_emu", "top_emu", "width_emu", "height_emu",
                            "hyperlink_left", "hyperlink_top", "hyperlink_size"):
                cell.alignment = centre_align
            else:
                cell.alignment = left_align

        # Apply function dropdown
        func_col = COLUMNS.index("function") + 1
        func_dv.add(ws.cell(row=excel_row, column=func_col))

        # Apply enabled dropdown
        enabled_col = COLUMNS.index("enabled") + 1
        enabled_dv.add(ws.cell(row=excel_row, column=enabled_col))

        # Apply scope dropdown
        scope_col = COLUMNS.index("scope") + 1
        scope_dv.add(ws.cell(row=excel_row, column=scope_col))

        # Per-row base_chart_name dropdown — constrained to the valid chart
        # refs for the row's data shape via the shared rule in dialog_support
        # (which itself falls back to all refs if the shape is unknown).
        if func == "insert_chart":
            cache_file = str(row.get("cache_file") or "").strip()
            if cache_file.lower() == "none":
                cache_file = ""
            converts_to_metrics = bool(str(row.get("metric_periods") or "").strip())
            chart_refs = get_valid_chart_refs_for_cache_file(
                cache_file, manifest or {}, converts_to_metrics, custom_chart_rows=custom_chart_rows
            )

            if chart_refs:
                ref_formula = '"' + ",".join(chart_refs) + '"'
                chart_dv = DataValidation(
                    type="list",
                    formula1=ref_formula,
                    allow_blank=True,
                    showDropDown=False,
                )
                ws.add_data_validation(chart_dv)
                ctr_col = COLUMNS.index("base_chart_name") + 1
                chart_dv.add(ws.cell(row=excel_row, column=ctr_col))

            # Per-row period dropdowns, only for a TimeSeries cache_file
            # with a known period list. All three columns validate against
            # the same hidden-sheet range, built once above.
            #
            # Excel list validation is single-value only, so the dropdown
            # offers one period at a time. A cell already holding a
            # '^'-delimited metric_periods value is not blocked by it.
            period_dv = period_dv_by_cache_file.get(row_cache_file)
            if period_dv is not None:
                for period_col_name in ("start_period", "end_period", "metric_periods"):
                    period_col = COLUMNS.index(period_col_name) + 1
                    period_dv.add(ws.cell(row=excel_row, column=period_col))

        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path
